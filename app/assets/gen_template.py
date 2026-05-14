"""
Gera o template.xlsx oficial do App Iter.

Schema novo de 15 colunas — ordem espelha a tela SACI (bloco "Dados do vôo"
seguido de "Tempo de vôo"). A coluna FUNCAO tem dropdown de validação com as
7 opções do SACI; CURSO_COMERCIAL tem dropdown Sim/Não.

Saída: app/assets/template.xlsx
Rodar: `python -m app.assets.gen_template`
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.core.config import FUNCOES_VALIDAS


# ── Paleta (mesma do app — preto + bege quente) ──────────────────────────────
COR_HEADER_BG       = "0B0B0E"  # preto Iter
COR_HEADER_FG       = "F5ECD8"  # bege quente
COR_PRIMARY         = "2B6BFF"  # indigo (opcional)
COR_BORDER          = "262630"  # cinza escuro
COR_GROUP_BG        = "1A1A22"  # preto suave (header de grupo)
COR_OBRIG_TEXT      = "F5ECD8"  # bege
COR_OPCIONAL_TEXT   = "C9D2E0"  # cinza claro

# ── Schema (ordem espelha a tela SACI) ───────────────────────────────────────
# (nome, descrição_tooltip, grupo, obrigatória?)
COLUNAS = [
    # Bloco "Dados do vôo"
    ("DATA",            "DD/MM/AAAA", "Dados do vôo", True),
    ("POUSOS",          "Inteiro 1-99", "Dados do vôo", True),
    ("FUNCAO",          "Escolha uma opção do dropdown", "Dados do vôo", True),
    ("ANAC_ALUNO",      "Código ANAC do aluno (8 dígitos). Obrigatório se "
                        "Função = Instrutor Voo ou Instrutor de voo em solo. "
                        "Vazio nas demais Funções.", "Dados do vôo", False),
    ("CURSO_COMERCIAL", "Sim/Não. Marque Sim apenas se Função = Piloto em "
                        "Comando E o voo foi dentro de curso de piloto "
                        "comercial aprovado pela ANAC com instrutor a bordo.",
                        "Dados do vôo", False),
    ("OBSERVACOES",     "Texto livre opcional", "Dados do vôo", False),
    ("MILHAS_NAV",      "Milhas náuticas (inteiro)", "Dados do vôo", False),
    # Bloco "Tempo de vôo"
    ("MATRICULA",       "Matrícula da aeronave (até 5 chars, ex.: PTBIC)",
                        "Tempo de vôo", True),
    ("ORIGEM",          "Aeródromo de origem — ICAO 4 letras (ex.: SBSP)",
                        "Tempo de vôo", True),
    ("DESTINO",         "Aeródromo de destino — ICAO 4 letras",
                        "Tempo de vôo", True),
    ("DIURNO",          "HH:MM. Pelo menos um de Diurno ou Noturno deve estar "
                        "preenchido.", "Tempo de vôo", False),
    ("NOTURNO",         "HH:MM. Pelo menos um de Diurno ou Noturno deve estar "
                        "preenchido.", "Tempo de vôo", False),
    ("NAVEGACAO",       "HH:MM — tempo de navegação", "Tempo de vôo", False),
    ("INSTRUMENTO",     "HH:MM — Instrumento Real (IFR efetivo)",
                        "Tempo de vôo", False),
    ("SOB_CAPOTA",      "HH:MM — instrução IFR simulada sob capota",
                        "Tempo de vôo", False),
]

# Larguras por coluna — valores aprovados pelo usuário (Excel auto-fit em
# 2026-05-14, refletindo o tamanho real do conteúdo de cada coluna).
LARGURAS = {
    "DATA": 12.0,
    "POUSOS": 11.66,
    "FUNCAO": 26.0,
    "ANAC_ALUNO": 16.22,
    "CURSO_COMERCIAL": 21.0,
    "OBSERVACOES": 30.0,
    "MILHAS_NAV": 14.33,
    "MATRICULA": 14.11,
    "ORIGEM": 9.0,
    "DESTINO": 13.22,
    "DIURNO": 9.0,
    "NOTURNO": 14.44,
    "NAVEGACAO": 15.89,
    "INSTRUMENTO": 16.66,
    "SOB_CAPOTA": 15.11,
}

# Linha 3 — exemplo enxuto. ICAOs ficam como "ZZZZ" pra deixar claro que é
# placeholder (o usuário substitui pelos códigos reais). Campos opcionais
# ficam vazios pra não dar impressão de obrigatoriedade.
EXEMPLO = {
    "DATA": "12/03/2026", "POUSOS": "1", "FUNCAO": "Piloto em Comando",
    "ANAC_ALUNO": "", "CURSO_COMERCIAL": "Não", "OBSERVACOES": "Translado",
    "MILHAS_NAV": "50", "MATRICULA": "PTBIC", "ORIGEM": "ZZZZ",
    "DESTINO": "ZZZZ", "DIURNO": "00:45", "NOTURNO": "00:30",
    "NAVEGACAO": "00:45", "INSTRUMENTO": "", "SOB_CAPOTA": "",
}


def _border(cor: str = COR_BORDER) -> Border:
    side = Side(style="thin", color=cor)
    return Border(left=side, right=side, top=side, bottom=side)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Voos"

    nomes = [c[0] for c in COLUNAS]

    # ── Linha 1: grupos visuais ("Dados do vôo" / "Tempo de vôo") ───────────
    # Acha as faixas de colunas pra cada grupo e faz merge
    grupo_atual = None
    inicio = 1
    for col_idx, (_nome, _tip, grupo, _ob) in enumerate(COLUNAS, start=1):
        if grupo != grupo_atual:
            if grupo_atual is not None:
                _aplicar_header_grupo(ws, inicio, col_idx - 1, grupo_atual)
            inicio = col_idx
            grupo_atual = grupo
    # Último grupo
    _aplicar_header_grupo(ws, inicio, len(COLUNAS), grupo_atual)
    ws.row_dimensions[1].height = 22

    # ── Linha 2: cabeçalhos das colunas ─────────────────────────────────────
    header_font = Font(name="Segoe UI", size=11, bold=True, color=COR_HEADER_FG)
    header_fill = PatternFill(
        start_color=COR_HEADER_BG, end_color=COR_HEADER_BG, fill_type="solid",
    )
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, (nome, tooltip, _grp, obrig) in enumerate(COLUNAS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=nome)
        cell.font = Font(
            name="Segoe UI", size=11, bold=True,
            color=COR_OBRIG_TEXT if obrig else COR_OPCIONAL_TEXT,
        )
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = _border()
        cell.comment = Comment(
            f"{tooltip}\n\n{'[Obrigatório]' if obrig else '[Opcional]'}",
            "App Iter",
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = LARGURAS[nome]
    ws.row_dimensions[2].height = 26

    # ── Linha 3: exemplo preenchido ─────────────────────────────────────────
    body_font = Font(name="Segoe UI", size=11)
    align_left = Alignment(horizontal="left", vertical="center")
    align_center_body = Alignment(horizontal="center", vertical="center")
    for col_idx, nome in enumerate(nomes, start=1):
        val = EXEMPLO.get(nome, "")
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.font = body_font
        cell.alignment = align_left if nome == "OBSERVACOES" else align_center_body
        cell.border = _border()

    # ── Data validations ────────────────────────────────────────────────────
    # FUNCAO (coluna 3 = C): dropdown com as 7 opções
    funcao_letter = get_column_letter(nomes.index("FUNCAO") + 1)
    funcao_formula = '"' + ",".join(FUNCOES_VALIDAS) + '"'
    dv_funcao = DataValidation(
        type="list", formula1=funcao_formula, allow_blank=True,
        showDropDown=False,  # False = mostra a setinha
    )
    dv_funcao.error = (
        "Escolha uma das opções: Instrutor Voo, Piloto em Comando, "
        "Piloto em Instrução, Instrutor de voo em solo, Co-Piloto Single Pilot, "
        "Co-Piloto Single Pilot com co-piloto por questão regulamentar, "
        "Co-Piloto Dual Pilot."
    )
    dv_funcao.errorTitle = "Função inválida"
    dv_funcao.prompt = "Clique na seta e escolha uma opção"
    dv_funcao.promptTitle = "Função a bordo"
    ws.add_data_validation(dv_funcao)
    dv_funcao.add(f"{funcao_letter}3:{funcao_letter}1000")

    # CURSO_COMERCIAL (coluna E): Sim/Não
    curso_letter = get_column_letter(nomes.index("CURSO_COMERCIAL") + 1)
    dv_curso = DataValidation(
        type="list", formula1='"Sim,Não"', allow_blank=True,
        showDropDown=False,
    )
    dv_curso.error = "Use 'Sim' ou 'Não' (vazio é tratado como 'Não')."
    dv_curso.errorTitle = "Valor inválido"
    dv_curso.prompt = "Sim apenas se Função = Piloto em Comando"
    dv_curso.promptTitle = "Curso comercial"
    ws.add_data_validation(dv_curso)
    dv_curso.add(f"{curso_letter}3:{curso_letter}1000")

    # ── Congelar linhas 1+2 pra rolar mantendo cabeçalho ────────────────────
    ws.freeze_panes = "A3"

    # ── Salvar ──────────────────────────────────────────────────────────────
    destino = Path(__file__).parent / "template.xlsx"
    wb.save(destino)
    print(f"template.xlsx salvo em {destino} ({len(COLUNAS)} colunas)")


def _aplicar_header_grupo(ws, col_inicio: int, col_fim: int, titulo: str) -> None:
    """Merge na linha 1 cobrindo as colunas do grupo, com fundo preto suave."""
    ws.merge_cells(start_row=1, start_column=col_inicio, end_row=1, end_column=col_fim)
    cell = ws.cell(row=1, column=col_inicio, value=titulo)
    cell.font = Font(
        name="Segoe UI", size=10, bold=True, color=COR_HEADER_FG, italic=True,
    )
    cell.fill = PatternFill(
        start_color=COR_GROUP_BG, end_color=COR_GROUP_BG, fill_type="solid",
    )
    cell.alignment = Alignment(horizontal="center", vertical="center")
    # Borda só na célula raiz da merge — visualmente fica OK
    cell.border = _border()


if __name__ == "__main__":
    main()
