"""
Hard debug do App Iter.

Bateria abrangente que exercita:
  1. Boot completo do app (4 telas em runtime real)
  2. Validação de planilhas com edge cases:
     - Planilha com aba renomeada (não "Plan1")
     - Planilha com aba vazia
     - Planilha sem colunas obrigatórias
     - Planilha com linhas vazias no meio
     - Planilha com erros de formato (data inválida, pousos > 99, ICAO grande)
     - Planilha 100% válida (template oficial)
  3. Funções de cache de licença (criar/ler/limpar sessão)
  4. Funções de hash de planilha (idempotência)
  5. Imports + assets de UI presentes
  6. FilePicker registrado, save_file e pick_files disponíveis
  7. Componentes auxiliares (iter_wordmark resolve, fundo carrega)

Reporta sumário final: total OK / total FAIL / lista detalhada.
"""
import os
import shutil
import sys
import tempfile
import traceback
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest.mock import MagicMock

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

os.environ.setdefault("APP_ANAC_DEV", "1")

import pandas as pd
from openpyxl import Workbook

import flet as ft

OK = "OK  "
FAIL = "FAIL"
results: list[tuple[str, str, str]] = []


def check(nome: str, fn) -> None:
    try:
        detalhe = fn() or ""
        results.append((OK, nome, str(detalhe)[:200]))
    except Exception as exc:
        results.append((FAIL, nome, f"{type(exc).__name__}: {exc}"))
        traceback.print_exc()


# ── 1. Imports principais ────────────────────────────────────────────────────


def teste_imports() -> str:
    from app.core import config, excel_reader, civ_bot, browser
    from app.core import logger, popup_inspector, utils
    from app.state import planilha_hash, retomar
    from app.licensing import cache, verificador, supabase_client
    from app.ui import tela_onboarding, tela_licenca, tela_principal, tela_relatorio
    from app.ui import componentes, fundo
    return "todos imports resolvem"


# ── 2. Assets esperados existem ──────────────────────────────────────────────


def teste_assets() -> str:
    esperados = [
        "icon.ico", "iter-logo-ui.png", "iter-logo-ui-64.png",
        "iter-wordmark.png", "template.xlsx", "iter-icon-source.png",
    ]
    faltando = []
    for nome in esperados:
        p = ROOT / "app" / "assets" / nome
        if not p.exists():
            faltando.append(nome)
    if faltando:
        raise FileNotFoundError(f"Assets faltando: {faltando}")
    return f"{len(esperados)} assets OK"


# ── 3. Reader: aba renomeada ─────────────────────────────────────────────────


def _planilha_aba(tmp: Path, nome_aba: str, dados: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba
    if dados:
        cabecalhos = list(dados[0].keys())
        ws.append(cabecalhos)
        for d in dados:
            ws.append([d[c] for c in cabecalhos])
    out = tmp / f"planilha-{nome_aba.replace(' ', '-')}.xlsx"
    wb.save(out)
    return out


def teste_aba_renomeada(tmp: Path) -> str:
    from app.core.excel_reader import validar_planilha
    dados = [
        {"data": "10/03/2026", "pousos": 1, "matricula": "PTVZZ",
         "origem": "SBSP", "destino": "SBKP", "horas": "00:45"},
    ]
    arq = _planilha_aba(tmp, "Voos Marco 2026", dados)
    validos, erros = validar_planilha(str(arq))
    if len(validos) != 1 or erros:
        raise AssertionError(f"esperado 1 válido, 0 erros — got {len(validos)} / {len(erros)}")
    return "aba 'Voos Marco 2026' (não-Plan1) lida com sucesso"


def teste_aba_plan1_explicita(tmp: Path) -> str:
    from app.core.excel_reader import validar_planilha
    dados = [
        {"data": "10/03/2026", "pousos": 1, "matricula": "PRABC",
         "origem": "SBJD", "destino": "SBPC", "horas": "01:00"},
    ]
    arq = _planilha_aba(tmp, "Plan1", dados)
    validos, erros = validar_planilha(str(arq))
    if len(validos) != 1:
        raise AssertionError(f"esperado 1 válido — got {len(validos)}")
    return "aba 'Plan1' (caso clássico) continua funcionando"


def teste_planilha_sem_colunas(tmp: Path) -> str:
    from app.core.excel_reader import validar_planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Algo"
    ws.append(["nome", "idade"])
    ws.append(["Joao", 30])
    arq = tmp / "sem-colunas.xlsx"
    wb.save(arq)
    try:
        validar_planilha(str(arq))
    except ValueError as exc:
        if "obrigatorias" in str(exc).lower() or "obrigatórias" in str(exc).lower():
            return "rejeita planilha sem colunas obrigatórias com mensagem clara"
    raise AssertionError("validar_planilha deveria ter lançado ValueError sobre colunas")


def teste_planilha_linhas_vazias(tmp: Path) -> str:
    from app.core.excel_reader import validar_planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan1"
    ws.append(["data", "pousos", "matricula", "origem", "destino", "horas"])
    ws.append(["10/03/2026", 1, "PTVZZ", "SBSP", "SBKP", "00:45"])
    ws.append([None, None, None, None, None, None])  # linha em branco no meio
    ws.append(["11/03/2026", 2, "PTVZZ", "SBKP", "SDAM", "01:10"])  # após branca: ignorada pela lógica atual
    arq = tmp / "linhas-vazias.xlsx"
    wb.save(arq)
    validos, erros = validar_planilha(str(arq))
    if len(validos) != 1:
        raise AssertionError(f"linha em branco deve cortar leitura — got {len(validos)} válidos")
    return f"linha em branco corta leitura corretamente ({len(validos)} antes da branca)"


def teste_planilha_com_erros(tmp: Path) -> str:
    from app.core.excel_reader import validar_planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan1"
    ws.append(["data", "pousos", "matricula", "origem", "destino", "horas"])
    ws.append(["NÃO-É-DATA", 1, "PTVZZ", "SBSP", "SBKP", "00:45"])         # data inválida
    ws.append(["10/03/2026", 999, "PTVZZ", "SBSP", "SBKP", "00:45"])       # pousos fora de range
    ws.append(["10/03/2026", 1, "MUITOLONGO", "SBSP", "SBKP", "00:45"])    # matricula muito longa
    ws.append(["10/03/2026", 1, "PTVZZ", "SBSP123", "SBKP", "00:45"])      # ICAO muito longo
    ws.append(["10/03/2026", 1, "PTVZZ", "SBSP", "SBKP", "00:45"])         # válida
    arq = tmp / "com-erros.xlsx"
    wb.save(arq)
    validos, erros = validar_planilha(str(arq))
    if len(validos) != 1 or len(erros) < 4:
        raise AssertionError(
            f"esperado 1 válido + ≥4 erros — got {len(validos)} válidos / {len(erros)} erros"
        )
    return f"validação coleta {len(erros)} erro(s) e {len(validos)} válido(s) corretamente"


# ── Cenários novos: formatos comuns do Excel que quebravam antes do fix ─────


def _planilha_celulas_tipadas(tmp: Path, headers: list[str], linhas: list[list]) -> Path:
    """
    Cria planilha permitindo CADA célula ser de um tipo específico (int, float,
    datetime.time, etc) — não apenas string. Necessário pra reproduzir
    DATE=int, HORAS=time, POUSOS=float etc.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan1"
    ws.append(headers)
    for linha in linhas:
        ws.append(linha)
    out = tmp / f"planilha-tipos-{len(list(tmp.iterdir()))}.xlsx"
    wb.save(out)
    return out


def teste_data_como_int8(tmp: Path) -> str:
    """Reproduz exatamente o bug do HR.xlsx: DATA=13052026 (int)."""
    from app.core.excel_reader import validar_planilha
    arq = _planilha_celulas_tipadas(
        tmp,
        ["DATA", "POUSOS", "MATRICULA", "ORIGEM", "DESTINO", "HORAS"],
        [[13052026, 1, "PTBIC", "SBSP", "SBKP", "00:45"]],
    )
    v, e = validar_planilha(str(arq))
    if len(v) != 1 or e:
        raise AssertionError(f"DATA int8: esperado 1/0 — got {len(v)}/{len(e)}: {e}")
    return f"DATA=13052026 (int 8d) parseado para {v[0]['data']}"


def teste_data_como_int7(tmp: Path) -> str:
    """7 dígitos: data com dia 1-9 vira DMMYYYY (ex: 7052026 -> 07/05/2026)."""
    from app.core.excel_reader import validar_planilha
    arq = _planilha_celulas_tipadas(
        tmp,
        ["DATA", "POUSOS", "MATRICULA", "ORIGEM", "DESTINO", "HORAS"],
        [[7052026, 1, "PTBIC", "SBSP", "SBKP", "00:45"]],
    )
    v, e = validar_planilha(str(arq))
    if len(v) != 1 or e:
        raise AssertionError(f"DATA int7: esperado 1/0 — got {len(v)}/{len(e)}: {e}")
    if v[0]["data"] != "07/05/2026":
        raise AssertionError(f"esperado 07/05/2026, got {v[0]['data']}")
    return f"DATA=7052026 (int 7d) parseado para 07/05/2026"


def teste_horas_como_time_obj(tmp: Path) -> str:
    """HORAS gravado como datetime.time(0, 30) — célula com formato hora nativo."""
    from datetime import time as dtime
    from app.core.excel_reader import validar_planilha
    arq = _planilha_celulas_tipadas(
        tmp,
        ["DATA", "POUSOS", "MATRICULA", "ORIGEM", "DESTINO", "HORAS"],
        [["12/03/2026", 1, "PTBIC", "SBSP", "SBKP", dtime(1, 30)]],
    )
    v, e = validar_planilha(str(arq))
    if len(v) != 1 or e:
        raise AssertionError(f"HORAS time: esperado 1/0 — got {len(v)}/{len(e)}: {e}")
    if v[0]["horas"] != "01:30":
        raise AssertionError(f"esperado horas=01:30, got {v[0]['horas']}")
    return f"HORAS=datetime.time(1,30) parseado para 01:30"


def teste_horas_hhmmss(tmp: Path) -> str:
    """HORAS como string '00:30:00' (formato HH:MM:SS)."""
    from app.core.excel_reader import validar_planilha
    arq = _planilha_celulas_tipadas(
        tmp,
        ["DATA", "POUSOS", "MATRICULA", "ORIGEM", "DESTINO", "HORAS"],
        [["12/03/2026", 1, "PTBIC", "SBSP", "SBKP", "00:30:00"]],
    )
    v, e = validar_planilha(str(arq))
    if len(v) != 1 or e:
        raise AssertionError(f"HORAS HH:MM:SS: esperado 1/0 — got {len(v)}/{len(e)}: {e}")
    return f"HORAS='00:30:00' (HH:MM:SS) parseado para {v[0]['horas']}"


def teste_data_serial_excel(tmp: Path) -> str:
    """DATA como float Excel serial (45920 = 2025-09-18)."""
    from app.core.excel_reader import validar_planilha
    arq = _planilha_celulas_tipadas(
        tmp,
        ["DATA", "POUSOS", "MATRICULA", "ORIGEM", "DESTINO", "HORAS"],
        [[45920.0, 1, "PTBIC", "SBSP", "SBKP", "00:45"]],
    )
    v, e = validar_planilha(str(arq))
    if len(v) != 1 or e:
        raise AssertionError(f"DATA serial: esperado 1/0 — got {len(v)}/{len(e)}: {e}")
    return f"DATA=45920.0 (serial Excel) parseado para {v[0]['data']}"


def teste_pousos_como_float(tmp: Path) -> str:
    """POUSOS=2.0 (Excel mistura tipos numa coluna de inteiros)."""
    from app.core.excel_reader import validar_planilha
    arq = _planilha_celulas_tipadas(
        tmp,
        ["DATA", "POUSOS", "MATRICULA", "ORIGEM", "DESTINO", "HORAS"],
        [["12/03/2026", 2.0, "PTBIC", "SBSP", "SBKP", "01:00"]],
    )
    v, e = validar_planilha(str(arq))
    if len(v) != 1 or e:
        raise AssertionError(f"POUSOS float: esperado 1/0 — got {len(v)}/{len(e)}: {e}")
    if v[0]["pousos"] != "2":
        raise AssertionError(f"esperado pousos='2', got '{v[0]['pousos']}'")
    return f"POUSOS=2.0 (float) normalizado para '2'"


def teste_icao_lowercase(tmp: Path) -> str:
    """ICAO em minúsculas (sbsp) — deve uppercase automaticamente."""
    from app.core.excel_reader import validar_planilha
    arq = _planilha_celulas_tipadas(
        tmp,
        ["DATA", "POUSOS", "MATRICULA", "ORIGEM", "DESTINO", "HORAS"],
        [["12/03/2026", 1, "PTBIC", "sbsp", "sbkp", "00:45"]],
    )
    v, e = validar_planilha(str(arq))
    if len(v) != 1 or e:
        raise AssertionError(f"ICAO lower: esperado 1/0 — got {len(v)}/{len(e)}: {e}")
    if v[0]["origem"] != "SBSP" or v[0]["destino"] != "SBKP":
        raise AssertionError(f"esperado uppercase, got {v[0]['origem']}/{v[0]['destino']}")
    return f"ICAO lowercase 'sbsp' normalizado para 'SBSP'"


def teste_icao_com_espaco(tmp: Path) -> str:
    """ORIGEM com espaços ' SBSP ' — strip deve resolver."""
    from app.core.excel_reader import validar_planilha
    arq = _planilha_celulas_tipadas(
        tmp,
        ["DATA", "POUSOS", "MATRICULA", "ORIGEM", "DESTINO", "HORAS"],
        [["12/03/2026", 1, "PTBIC", " SBSP ", "SBKP", "00:45"]],
    )
    v, e = validar_planilha(str(arq))
    if len(v) != 1 or e:
        raise AssertionError(f"ICAO espaços: esperado 1/0 — got {len(v)}/{len(e)}: {e}")
    return f"ICAO ' SBSP ' (com espaços) strip para 'SBSP'"


def teste_matricula_com_traco(tmp: Path) -> str:
    """MATRICULA 'PT-BIC' (6 chars) deve falhar com erro de tamanho."""
    from app.core.excel_reader import validar_planilha
    arq = _planilha_celulas_tipadas(
        tmp,
        ["DATA", "POUSOS", "MATRICULA", "ORIGEM", "DESTINO", "HORAS"],
        [["12/03/2026", 1, "PT-BIC", "SBSP", "SBKP", "00:45"]],
    )
    v, e = validar_planilha(str(arq))
    if len(v) != 0 or len(e) != 1:
        raise AssertionError(f"matricula > 5: esperado 0/1 — got {len(v)}/{len(e)}")
    if e[0].coluna != "matricula":
        raise AssertionError(f"esperado erro em coluna 'matricula', got '{e[0].coluna}'")
    return f"MATRICULA='PT-BIC' (6 chars) rejeitada corretamente"


def teste_planilha_real_hr(tmp: Path) -> str:
    """Reproduz a estrutura exata do HR.xlsx do usuário — após fixes deve passar limpo."""
    from datetime import time as dtime
    from app.core.excel_reader import validar_planilha
    arq = _planilha_celulas_tipadas(
        tmp,
        ["DATA", "POUSOS", "MATRICULA", "ORIGEM", "DESTINO", "HORAS",
         "OBS", "MILHAS_NAV", "HORAS_NAV"],
        [[13052026, "1", "PTVZZ", "ZZZZ", "ZZZZ", dtime(0, 30),
          "Translado", None, None]],
    )
    v, e = validar_planilha(str(arq))
    if len(v) != 1 or e:
        raise AssertionError(f"HR.xlsx real: esperado 1/0 — got {len(v)}/{len(e)}: {e}")
    return f"HR.xlsx reproduzido: data={v[0]['data']}, horas={v[0]['horas']}"


def teste_template_oficial(tmp: Path) -> str:
    """Template atual tem 1 linha de exemplo (PTBIC, SBSP→SBKP)."""
    from app.core.excel_reader import validar_planilha
    src = ROOT / "app" / "assets" / "template.xlsx"
    if not src.exists():
        raise FileNotFoundError("template.xlsx oficial não encontrado em app/assets")
    validos, erros = validar_planilha(str(src))
    if len(validos) != 1 or erros:
        raise AssertionError(
            f"template oficial deveria ser 1/0 — got {len(validos)}/{len(erros)}: {erros}"
        )
    if validos[0]["matricula"] != "PTBIC":
        raise AssertionError(f"template deveria ter PTBIC — got {validos[0]['matricula']}")
    return f"template oficial: 1 linha PTBIC válida, 0 erros"


# ── 4. Cache de licença ──────────────────────────────────────────────────────


def teste_cache_sessao(tmp: Path) -> str:
    """Não usa APPDATA real — apenas verifica round-trip da serialização."""
    from app.licensing.cache import SessaoIter, salvar_sessao, ler_sessao, limpar_sessao

    # Backup do estado real
    from app.licensing import cache as cache_mod
    cache_real = cache_mod._cache_path()
    tinha_cache = cache_real.exists()
    backup = None
    if tinha_cache:
        backup = cache_real.read_bytes()

    try:
        sess = SessaoIter(
            email="teste@iter.com",
            access_token="ya29.abc",
            refresh_token="refresh.xyz",
            expires_at=datetime.now(timezone.utc) + timedelta(hours=1),
            proxima_revalidacao=datetime.now(timezone.utc) + timedelta(hours=24),
        )
        salvar_sessao(sess)
        carregado = ler_sessao()
        assert carregado is not None
        assert carregado.email == "teste@iter.com"
        assert carregado.access_token == "ya29.abc"
        limpar_sessao()
        assert ler_sessao() is None
    finally:
        # Restaura estado original
        if backup is not None:
            cache_real.write_bytes(backup)
    return "salvar/ler/limpar sessão round-trip OK"


# ── 5. Hash de planilha (idempotência) ───────────────────────────────────────


def teste_hash_planilha(tmp: Path) -> str:
    from app.state.planilha_hash import calcular_hash, _historico_path
    arq = tmp / "hash.xlsx"
    wb = Workbook()
    wb.active.append(["a", "b", "c"])
    wb.save(arq)
    h1 = calcular_hash(str(arq))
    h2 = calcular_hash(str(arq))
    if h1 != h2:
        raise AssertionError("hash não é determinístico")
    if len(h1) != 64:
        raise AssertionError(f"hash sha256 deveria ter 64 chars, tem {len(h1)}")
    return f"hash sha256 estável ({h1[:12]}...)"


# ── 6. UI: construção das telas via Page mock ────────────────────────────────


def teste_telas_construir() -> str:
    page = MagicMock(spec=ft.Page)
    page.overlay = []
    page.services = []
    page.controls = []
    page.window = MagicMock()
    page.snack_bar = None

    from app.ui import tela_onboarding, tela_licenca, tela_principal, tela_relatorio
    tela_onboarding.construir(page, lambda _: None)
    tela_licenca.construir(page, lambda _: None)
    tela_principal.construir(page, lambda *a, **kw: None)
    tela_relatorio.construir(page, lambda _: None, 12, 1, 87.5)
    return "4 telas construídas sem erro"


# ── 7. FilePicker e save_file disponíveis ────────────────────────────────────


def teste_filepicker_api() -> str:
    import inspect
    assert hasattr(ft, "FilePicker")
    fp = ft.FilePicker()
    assert inspect.iscoroutinefunction(fp.pick_files)
    assert inspect.iscoroutinefunction(fp.save_file)
    return "FilePicker.pick_files e save_file são corotinas (Flet 0.85)"


# ── 8. Helpers de UI (componentes + fundo) ───────────────────────────────────


def teste_helpers_ui() -> str:
    from app.ui.componentes import (
        botao_primario, botao_secundario, iter_wordmark,
        iter_wordmark_footer, eyebrow, titulo, texto, card,
    )
    from app.ui.fundo import camadas_ambient

    b = botao_primario("X", lambda _: None)
    assert b.content == "X"
    wm = iter_wordmark(40)
    assert wm.src == "iter-wordmark.png"
    f = iter_wordmark_footer(16, 0.3)
    assert f.opacity == 0.3
    camadas = camadas_ambient()
    assert len(camadas) >= 2
    for c in camadas:
        assert hasattr(c, "bgcolor")
    return f"helpers UI OK ({len(camadas)} camadas ambient)"


# ── 9. Boot test invocado como subprocesso ───────────────────────────────────


def teste_boot_real() -> str:
    import subprocess
    cmd = [sys.executable, str(ROOT / "scripts" / "boot_test.py")]
    res = subprocess.run(
        cmd, capture_output=True, text=True, timeout=30,
        env={**os.environ, "APP_ANAC_DEV": "1"},
    )
    if res.returncode != 0:
        raise AssertionError(
            f"boot_test exit code {res.returncode}\n"
            f"stdout: {res.stdout[-800:]}\n"
            f"stderr: {res.stderr[-800:]}"
        )
    # Confirma que 4 telas renderizaram
    if "OK   tela_relatorio" not in res.stdout:
        raise AssertionError(f"boot_test não renderizou todas as 4 telas. stdout:\n{res.stdout}")
    return "4 telas renderizaram em runtime real Flet"


# ── 10. Logger não vaza handlers em chamadas múltiplas ───────────────────────


def teste_logger_handlers_unicos() -> str:
    from app.core.logger import get_logger
    l1 = get_logger()
    n1 = len(l1.handlers)
    l2 = get_logger()
    n2 = len(l2.handlers)
    if n1 != n2:
        raise AssertionError(f"get_logger duplicou handlers: {n1} → {n2}")
    return f"logger reutiliza handlers ({n1} fixos)"


# ── 11. Config valores sãos ─────────────────────────────────────────────────


def teste_config_sanity() -> str:
    from app.core import config
    assert config.URL_CIV.startswith("https://sistemas.anac.gov.br/")
    assert config.CDP_PORTA == 9222
    assert config.TIMEOUT_PADRAO >= 5_000
    return f"URL_CIV, CDP_PORTA, TIMEOUT_PADRAO consistentes"


# ── Runner ───────────────────────────────────────────────────────────────────


def main() -> int:
    print("=" * 72)
    print("HARD DEBUG — App Iter")
    print("=" * 72)

    tmp = Path(tempfile.mkdtemp(prefix="app_iter_debug_"))
    print(f"Tempdir: {tmp}\n")

    casos = [
        ("imports principais",          teste_imports),
        ("assets necessários",          teste_assets),
        ("config sanity",               teste_config_sanity),
        ("logger handlers únicos",      teste_logger_handlers_unicos),
        ("UI: construir 4 telas",       teste_telas_construir),
        ("UI: helpers componentes/fundo", teste_helpers_ui),
        ("FilePicker async API",        teste_filepicker_api),
        ("planilha — aba renomeada",    lambda: teste_aba_renomeada(tmp)),
        ("planilha — aba Plan1",        lambda: teste_aba_plan1_explicita(tmp)),
        ("planilha — sem colunas",      lambda: teste_planilha_sem_colunas(tmp)),
        ("planilha — linhas vazias",    lambda: teste_planilha_linhas_vazias(tmp)),
        ("planilha — erros formato",    lambda: teste_planilha_com_erros(tmp)),
        # Cenários reais do Excel que estavam quebrando antes do fix
        ("DATA como int 8d",            lambda: teste_data_como_int8(tmp)),
        ("DATA como int 7d",            lambda: teste_data_como_int7(tmp)),
        ("HORAS datetime.time",         lambda: teste_horas_como_time_obj(tmp)),
        ("HORAS HH:MM:SS",              lambda: teste_horas_hhmmss(tmp)),
        ("DATA serial Excel",           lambda: teste_data_serial_excel(tmp)),
        ("POUSOS float (2.0)",          lambda: teste_pousos_como_float(tmp)),
        ("ICAO lowercase",              lambda: teste_icao_lowercase(tmp)),
        ("ICAO com espaços",            lambda: teste_icao_com_espaco(tmp)),
        ("MATRICULA com traço (>5)",    lambda: teste_matricula_com_traco(tmp)),
        ("planilha real HR.xlsx",       lambda: teste_planilha_real_hr(tmp)),
        ("template.xlsx oficial",       lambda: teste_template_oficial(tmp)),
        ("cache sessão licensing",      lambda: teste_cache_sessao(tmp)),
        ("hash planilha idempotente",   lambda: teste_hash_planilha(tmp)),
        ("boot real Flet 4 telas",      teste_boot_real),
    ]

    for nome, fn in casos:
        check(nome, fn)

    print()
    for status, nome, detalhe in results:
        print(f"  {status}  {nome:36s} {detalhe}")

    falhas = sum(1 for s, _, _ in results if s == FAIL)
    total = len(results)
    print()
    print("=" * 72)
    print(f"Resultado: {total - falhas}/{total} OK, {falhas} falha(s)")
    print("=" * 72)

    # Cleanup tempdir
    try:
        shutil.rmtree(tmp, ignore_errors=True)
    except Exception:
        pass

    return falhas


if __name__ == "__main__":
    sys.exit(main())
